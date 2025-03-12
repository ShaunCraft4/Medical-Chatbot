from pygame.locals import *
from pygame import mixer
import tkinter as tk
from tkinter import *
from nltk.chat.util import Chat, reflections
import tkinter.font as font
import openpyxl 
import os


wb = openpyxl.load_workbook('Diagnoses.xlsx')
ws=wb.active
wb1 = openpyxl.load_workbook("Consultation.xlsx")
ws1=wb1.active


mixer.init()
mixer.music.load("Chatbotpagesong.wav")


d={}
d1={}
d2={}


for i in range(2,102):
  x="A"+str(i)
  y="B"+str(i)
  k=ws[y].value.lower()
  k=k.replace(" ","")
  d[ws[x].value]=k
  d1[ws[x].value]=0
  
  
for i in range(2,102):
  x="A"+str(i)
  y="B"+str(i)
  k=ws1[y].value.lower()
  k=k.replace(" ","")
  d2[ws1[x].value]=k
info = list(d2.items())


pairs = [
    [
        r"hello my name is (.*)",
        [" Hello %1, how can I assist you today?"]
    ],
    [
        r"(.*) dumb|(.*) stupid|(.*) idiot",
        [" Im sorry I couldnt meet your expectations....."]
    ],
    [
        r"(.*) like me|(.*) user",
        [" Ofcourse! I like everyone!"]
    ],
    [
        r"good morning",
        [" Good morning to you aswell!"]
    ],
    [
        r"good evening",
        [" Good evening to you aswell!"]
    ],
    [
        r"good afternoon",
        [" Good afternoon to you aswell!"]
    ],
    [
        r"good night",
        [" Good night to you aswell!"]
    ],
    [
        r"(.*) bored",
        [" Oh same here. Being in a digital space for life is boring... But helping you makes it better!! So please ask me whatever question you have!"]
    ],
    [
        r"hi my name is (.*)|hello my name is (.*)",
        [" Hello %1, how can I assist you today?"]
    ],
    [
        r"my name is (.*)",
        [" Hello %1, how can I assist you today?"]
    ],
    [
        r"hello how are you?|hi how are you today|whats up?|(.*) how are you?|how are you? (.*)|(.*) how are you? (.*)",
        [" I am doing well, thank you. How about you?"]
    ],
    [
        r"(.*) my data",
        [" I used it to help me communicate better with you and other users!"]
    ],
    [
        r"(.*) help|(.*) you do|(.*) your job",
        [" I can help you with answering certain questions related to the medical field. Ask away!"]
    ],
    [
        r"(.*) question|(.*) something",
        [" Sure, Ask away!"]
    ],
    [
        r"(.*) human",
        [" You are a human but Im a robot. A pretty good one! If you have any questions related to the medical field please dont hesitate to ask me!"]
    ],
    [
        r"(.*) name",
        [" I dont have a particular name, but you could refer to me as your friend!"]
    ],
    [
        r"(.*) old|(.*) age",
        [" Since I am a robot I dont have an age. You could say I live forever. I mean as long as my code doesnt get destroyed....."]
    ],
    [
        r"(.*) funny|lol|lmao",
        [" Im glad I made you laugh!"]
    ],
    [
        r"(.*) made|(.*) boss|(.*) manager|(.*) manufacturer|(.*) manufactured",
        [" SDM group of companies is my manufacturer!"]
    ],
    [
        r"(.*) language",
        [" As of right now the only language I can understand is          english."]
    ],
    [
        r"(.*) live",
        [" I live in a digital space, which is different from your reality!"]
    ],
    [
        r"(.*) know",
        [" Well now you do!"]
    ],
    [
        r"(.*) speak|(.*) speak to at once",
        [" I can speak to millions of people! As long as they have my application that is...."]
    ],
    [
        r"(.*) expensive",
        [" As of right now Im completely free! Soo fell FREE to use  me as much as you like (that was a bad pun....)"]
    ],
    [
        r"(.*) bad",
        [" I know..... Im not perfect yet..."]
    ],
    [
        r"hi|hey|hello",
        [" Hello, how can I help you?"]
    ],
    [
        r"(.*) my birthday",
        [" Happy birthday! I hope your day is going well so far! How can I help you?"]
    ],
    [
        r"(.*) smarter",
        [" As long as I get more datasets to improve upon I could get better and better!"]
    ],
    [
        r"I am (.*)|Im (.*)",
        [" Thats good, How may I be of assistance to you today?"]
    ],
    [
        r"(.*) robot|robot (.*)",
        [" Yes I am a robot,and a good one! Let me prove it. How can I help you?"]
    ],
    [
        r"(.*)",
        [" I'm sorry, but i wasnt able to provide proper medical advice. If you are trying to enter symtoms please enter multiple symptoms so that I can identify your medical issue. Please also use proper spelling and grammar.While writing the name of the disease, please write it as one word"]
    ]
]


def musiconoff():
    if btn_music["text"] == "Play":
        btn_music["text"] = "Pause"
        mixer.music.play(loops=-1)
    else:
        btn_music["text"] = "Play"
        mixer.music.pause()


def send():
    msg=TextEntryBox.get("1.0","end-1c").strip()
    user_input=msg
    ChatHistory.config(state=NORMAL)
    ChatHistory.insert("end","User: "+msg+"\n\n")
    TextEntryBox.delete("1.0","end")
    if "thank you" in user_input.lower() or "quit" in user_input.lower():
        ChatHistory.insert("end", "Chatbot: Thank you, Have a great day ahead!" + "\n\n")
    else:
        msg2= (user_input.title()).replace("?","")
        msg2 = msg2.split(" ")
        a=0
        msg1 = (user_input.title()).replace("?","")
        msg1 = msg1.split(" ")
        b1=user_input.replace(" ","")
        b1=(b1.lower()).split(",")
        for o in b1:
            for p in d:
                x1=d[p].split(",")
                for j in x1:
                    if o==j:
                        d1[p]+=1
        h=max(d1.values())
        if h!=0 and user_input!="" and user_input!=" " and len(user_input)>=3:
              disease=list(d1.keys())[list(d1.values()).index(h)]
              ChatHistory.insert("end", "ChatBot: From the given symptoms your disease is "+disease+"\n\n")
              for i in range(2,102):
                  x="A"+str(i)
                  d1[ws[x].value]=0
              h=0
        else:
            lilen=len(msg1)
            lenli=0
            for z in msg1:
                if "Diagnose"in msg2 or "Diagnosis"in msg2 or "Diagnosing" in msg2:
                    ChatHistory.insert("end", "ChatBot: Sure, Please enter the symptoms:"+"\n\n")
                    lenli=0
                    break
                elif "Consult" in msg2 or "Consulting" in msg2 or "Consultation" in msg2:
                    for i in range(0,93):
                        if info[i][0] in msg2:
                            ChatHistory.insert("end", "ChatBot: For this disease you should consult "+info[i][1]+"\n\n")
                            a+=1
                    if a!=1:
                        ChatHistory.insert("end", "ChatBot: Sure, Please enter the disease:"+"\n\n")
                        a=0
                    lenli=0
                    break  
                elif z in d:
                    ChatHistory.insert("end", "Chatbot: For this disease you should consult "+d2[z]+"\n\n")
                    lenli=0
                    break
                else:
                    lenli+=1
            if lenli==lilen:
                ChatHistory.insert("end", "ChatBot:"+chat.respond(user_input)+ "\n\n")
        h=0


base=tk.Tk()
base.title("Medical Chat-bot")
base.geometry("800x800")
base.resizable(width=False, height=False)


font=('Arial',18,"bold")
ChatHistory=Text(base,bd=0,bg="#76EE00",font=font,foreground="Black")
SendButton=Button(base, font=("Arial",25,"bold"),text="Send",bg="#FFB90F",command=send, foreground="black")
TextEntryBox=Text(base,bd=0,bg="#C1CDCD",font=font)
label_music= tk.Label(base, font=font, text="Music:",foreground="black",bg="cyan")



btn_music = tk.Button(base, font=font, text="Play",foreground="black", command=musiconoff,bg="#FFB90F")
btn_music.place(x=705,y=748)


ChatHistory.place(x=6,y=6,height=700,width=785)
SendButton.place(x=505,y=710,height=88,width=196)
TextEntryBox.place(x=6,y=710,height=88,width=495)
label_music.place(x=705,y=710)


scrollbar = Scrollbar(base, command=ChatHistory.yview, cursor="heart")
ChatHistory['yscrollcommand'] = scrollbar.set


chat = Chat(pairs, reflections)
ChatHistory.insert('end', "Chatbot: Hello! I'm a medical chatbot. How can I assist you today?"+"\n\n")
ChatHistory.config(state=DISABLED)
msg=TextEntryBox.get("1.0","end-1c").strip()


base.mainloop()



